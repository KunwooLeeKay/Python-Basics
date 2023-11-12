from tkinter import *
import tkinter.ttk as ttk
from tkinter import filedialog
from openpyxl import load_workbook 
import re
import math
from openpyxl import Workbook
import tkinter.messagebox as msg
import datetime
import os

root = Tk()
root.title("Kunwoo's Program1")
root.geometry("500x400")

def add_file():
    files = filedialog.askopenfilename(title = "파일을 선택하세요",\
         filetypes =(("상관없음", "*.*"), ("엑셀1", "*.xls"), ("엑셀2", "*.xlsx")),\
             initialdir = "C:/")
    file_name.delete(0, END)
    file_name.insert(END, files)

def path_sel():
    folder_selected = filedialog.askdirectory()
    if folder_selected == 'None':
        return
    path_name.delete(0, END)
    path_name.insert(0, folder_selected)

def start():
    try:
        # print(file_name.get())
        # wb = load_workbook(str(file_name)) # Entry 내의 값을 가져올때는 .get()을 해야한다.
        
        wb = load_workbook(file_name. get())
        ws = wb.active

        file_number = re.findall("\d+",file_name.get()) # re.findall을 이용해서 파일 내 숫자 추출
        date = file_number.pop() # 숫자중 날짜만 추출
        dates = date[6:8]

        multi = float(cmb_mul.get())

        drug_code = [] # 약물 코드 리스트
        drug_name = [] # 약물 이름 리스트
        current_stock = [] # 현재 재고량 리스트
        estimated_usage = [] # 예상 사용량 리스트 = 평균 일일 사용량 * 몇일치 발주 * 가중치
        how_many_days = (re.findall("\d+", days.get())). pop()
        box = [10,10,10,10,10,10,5,25,25,25,100,100,56,56,30,30,30,10,10,10] # 박스단위 리스트

        for y in range(4, 24):
            drug_code.append(ws["C{}". format(y)].value)
            drug_name.append(ws["D{}". format(y)].value)
            current_stock.append(ws["AE{}". format(y)].value)        
            estimated_usage.append(math.ceil(int(ws["AI{}". format(y)].value) \
                * int(how_many_days) * multi / int(dates)))
                # (당월 누적 사용량 / 오늘 날짜) * 몇일치 발주 * 가중치
                # 평균 일일 사용량 * 몇일치 발주 * 가중치
        
        order = [] # 필요주문량

        for x in range(0, 20):
            if estimated_usage[x] >= current_stock[x]:
                order.append(estimated_usage[x] - current_stock[x])
            else:
                order.append(0)
        print(order)
        
        box_order = [] # 박스단위 고려 주문량(박스단위)
        for z in range(0,20):
            box_order.append(math.ceil(order[z] / box[z]))

        print(box_order)

        ws["AJ3"] = "주문 필요 수량"
        ws["AK3"] = "박스 단위"

        ws["AM3"] = "발주 일수"
        ws["AN3"] = str(how_many_days) + "일치"
        ws["AM4"] = "가중치"
        ws["AN4"] = "X" + str(multi)


        for row in range(4,24):
            ws["AJ{}". format(row)] = order[row - 4]
            ws["AK{}". format(row)] = box_order[row - 4]

        wb.save(file_name.get())
        wb.close()

        wb2 = Workbook() # 새 워크북을 생성 : 엑셀 파일을 생성함
        ws = wb2.active # 현재 활성화된 sheet를 가져옴 -> ws에 저장

        ws.title = "마약발주서" # sheet의 이름을 변경
        ws["A1"] = "NO"
        ws["B1"] = "약품코드"
        ws["C1"] = "약품명"
        ws["D1"] = "주문 수량"
        ws["E1"] = "박스 단위"
        ws["F1"] = "주문 박스"

        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10

        for y in range(2, 22):
            ws["A{}". format(y)] = y-1
            ws["B{}". format(y)] = drug_code[y-2]
            ws["C{}". format(y)] = drug_name[y-2]
            ws["D{}". format(y)] = order[y-2]
            ws["E{}". format(y)] = box[y-2]
            ws["F{}". format(y)] = box_order[y-2]

        file_name2 = str(datetime.date.today()) + " 마약발주서.xlsx"
        dest_path = os.path.join(path_name.get(), file_name2)
        wb2.save(dest_path)

        wb2.close()

        msg.showinfo("알림", "작업이 완료되었습니다.")
    except Exception as err:
        msg.showerror("에러!",err)


frame_file = Frame(root)
frame_file. pack(fill = "x", padx = 5, pady = 5)

add_file = Button(frame_file, text = "파일 선택", command = add_file)
add_file. pack(side = "right", padx = 5, pady = 3)

# del_file = Button(frame_file, text = "", command = del_file)
# del_file. pack(side = "left")

frame_filename = Frame(root)
frame_filename. pack(fill = "x", padx = 5, pady = 5)

file_name = Entry(frame_filename)
file_name. pack(fill = "x", padx = 5, pady = 5, ipady = 2)


frame_option = LabelFrame(root, text = "옵션 선택")
frame_option. pack(fill = "x", padx = 3, pady = 5)

label_mul = Label(frame_option, text = "가중치", width = 8)
label_mul.pack(side = "left")

opt_mul = ["1.0", "1.2", "1.4", "1.6", "1.8", "2.0"]
cmb_mul = ttk.Combobox(frame_option, state = "readonly", values = opt_mul, width = 14)
cmb_mul.current(0)
cmb_mul.pack(side = "left", padx = 5, pady = 5)

label_days = Label(frame_option, text = "발주 일수")
label_days. pack(side = "left", padx = 5, pady = 5)

days = Entry(frame_option, width = 30)
days.pack(side = "left", padx = 5, pady = 5)
days.insert(0, "7")

# label_type = Label(frame_option, text = "발주 일수", width = 8)
# label_type.pack(side = "left")

# opt_type = ["7일", "사용자 정의"]
# cmb_type = ttk.Combobox(frame_option, values = opt_type, width = 14)
# cmb_type.current(0)
# cmb_type.pack(side = "left", padx = 5, pady = 5)

frame_path = LabelFrame(root, text = "경로 선택")
frame_path. pack(fill = "x")

path_name = Entry(frame_path)
path_name. pack(side = "left", fill = "x", expand = True, padx = 5, pady =5)

path_sel = Button(frame_path, text = "저장경로 선택", command = path_sel)
path_sel.pack(side = "right", padx = 5, pady = 5)

frame_action = Frame(root)
frame_action.pack(fill = 'x', ipady = 5, padx = 5, pady = 5)

start = Button(frame_action, text = "시작", padx = 5, pady = 5, command = start)
start. pack(side = "right")

end = Button(frame_action, text = "종료", padx = 5, pady = 5, command = root.quit)
end.pack(side = "right")

root.mainloop()