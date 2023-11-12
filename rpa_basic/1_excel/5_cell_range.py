from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active

# 1줄씩 데이터 넣기 : append
ws.append(["번호", "영어", "수학"]) # append에는 리스트, 튜플, 레인지같은 값들이 가능하다.

for i in range(1, 11): # 1번부터 10번까지의 학생에게 랜덤의 영어, 수학점수를 부여함
    ws.append([i, randrange(0,101), randrange(0,101)]) # append에 넣을 때 리스트 형태로 해줌

col_B = ws["B"] # 이렇게 하면 B컬럼 = 영어 만 가져올 수 있음
# print(col_B)
for cell in col_B:
    print(cell.value)

# 범위의 column 가져오기
col_range = ws["B:C"] # 영어, 수학 column함께 가지고 오기. 이때 슬라이싱과 조금 다르게 B,C를 모두 포함한 것을 가져온다.
# 이렇게 쓰면 튜플의 형태로 B1, B2, B3, ..., B11, C1, C2, ..., C11 의 값을 가져온다.

# for cell in col_range:
#     print(cell. value) 이렇게 쓰면 안됨.

for cols in col_range: # 이렇게 해서 첫번쨰 for 문에서 column에 대한 반복문
    for cell in cols: # 두번째에서 row에 대한 포문을 넣어줘야한다.
        print(cell.value)

# 범위의 row 가져오기
row_title = ws[1] # 첫번째 row만 가져오기
for cell in row_title:
    print(cell.value)

row_range = ws[2:6] # 2번쨰에서 6번째 줄을 가져오기
for rows in row_range:
    for cell in rows:
        print(cell.value, end = " ")
    print()

from openpyxl.utils.cell import coordinate_from_string # 셀의 좌표정보를 가져오게 해주는 라이브러리

for rows in row_range:
    for cell in rows:
        print(cell.coordinate, end = " ") # .coordinate로 좌표 정보 가져오기
        xy = coordinate_from_string(cell.coordinate)
         # 위에 import한 것을 사용하면 A,B,C와 같은 값과 1,2,3의 값을 분리해줌.
         # 만약 파일이 커져서 AZ250처럼 슬라이싱이 힘들어지면 사용한다. xy[0] 은 A,B,C,... xy[1]은 1,2,3,...이다.
        print(xy, end = " ")
        print(xy[0], end = "")
        print(xy[1], end = " ")
    print()

wb.save("sample.xlsx")