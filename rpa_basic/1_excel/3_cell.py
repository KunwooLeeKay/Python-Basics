from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "Kunwoo Sheet"

# A1 셀에 1이라는 값 입력
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"]) # A1 셀의 정보를 출력(값이 아님!)
print(ws["A1"].value) # 이렇게해야 값
print(ws["A10"].value) # 값이 없으면 None 출력함

print(ws.cell(row = 1, column = 1).value) # 이렇게 row,column으로도 접근 가능하다. = A1
print(ws.cell(row = 1, column = 2).value) # 이렇게하면 B1

c = ws.cell(column = 3, row = 1, value = 10) # 이렇게 쓰면 ws["C1"] = 10 과 같은 말이다.
print(c.value) # 위처럼 c에 셀의 값을 할당하고 c를 ws.cell(column = , row = ) 대신 쓸수도 있다.

from random import *
# 반복문을 이용해서 랜덤 숫자 채우기
for x in range(1,11):
    for y in range(1,11):
        ws.cell(row = x, column = y, value = randrange(0,101))

wb.save("sample.xlsx")
