from openpyxl import Workbook
wb = Workbook()

ws = wb.create_sheet() # 새로운 시트 생성 : 현재 실행중인 시트 뒤에 새로운 시트를 하나 만듬
ws.title = "MySheet"
ws.sheet_properties.tabColor = "ff66ff" # 시트 색상을 선택가능. 이때 ff66ff는 구글에 rgb 검색해서 값 찾아와야함

ws1 = wb.create_sheet("YourSheet") # 생성하면서 이름 생성
# 현재상황 : 기본 시트, mysheet, yoursheet 가 있음
ws2 = wb.create_sheet("NewSheet", 2) # 2번째 인덱스에 시트 생성
# => 기본, my, newsheet, yoursheet 가 됨

new_ws_var = wb["NewSheet"] # 이렇게 하면 앞서 만든 NewSheet를 변수 ws2가 아닌 새로운 변수에 할당해서 사용할 수 있다. 
# Dict형태로 불러온다

print(wb.sheetnames) # 모든 시트 정보를 출력

# Sheet 복사
new_ws_var["A1"] = "Test" # 이렇게 하면 A1 셀에 Test라고 입력
target = wb.copy_worksheet(new_ws_var)
target.title = "Copied Sheet"

wb.save("sample.xlsx")