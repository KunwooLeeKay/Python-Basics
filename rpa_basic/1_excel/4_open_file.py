from openpyxl import load_workbook # 파일 불러오기
wb = load_workbook("sample.xlsx")
ws = wb.active 

# cell 데이터 불러오기
for x in range(1,11):
    for y in range(1,11):
        print(ws.cell(row = x, column = y).value, end = " ")
    print()

# cell 객수를 모를 때
for x in range(1, ws.max_row + 1): 
    # ws.max_row/max_column을 하면 최대 행/렬의 값을 불러올 수 있다. 여기선 range써서 1 더함
    for y in range(1, ws.max_column + 1):
        print(ws.cell(row = x, column = y).value, end = " ")
    print()